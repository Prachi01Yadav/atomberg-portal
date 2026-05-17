import csv
import io
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.checkin import Quarter, QuarterlyCheckin
from app.models.cycle import PerformanceCycle
from app.models.goal import Goal, GoalStatus
from app.models.user import User, UserRole
from app.services.cycle_service import get_active_quarter
from app.services.goal_service import cap_score_for_display


async def build_achievement_rows(
    db: AsyncSession, cycle_id: UUID, department: str | None = None
) -> list[dict]:
    q = (
        select(Goal)
        .options(selectinload(Goal.employee), selectinload(Goal.checkins))
        .where(Goal.cycle_id == cycle_id)
    )
    if department:
        q = q.join(User, Goal.employee_id == User.id).where(User.department == department)

    goals = list((await db.execute(q)).scalars().all())
    rows: list[dict] = []

    for goal in goals:
        by_q = {c.quarter: c for c in goal.checkins}
        latest_score = None
        for qtr in (Quarter.Q4, Quarter.Q3, Quarter.Q2, Quarter.Q1):
            if qtr in by_q and by_q[qtr].computed_score is not None:
                latest_score = by_q[qtr].computed_score
                break

        weighted_total = 0.0
        if latest_score is not None:
            weighted_total = goal.weightage * cap_score_for_display(latest_score) / 100

        def actual_for(q: Quarter) -> str | float:
            if q not in by_q:
                return ""
            c = by_q[q]
            if goal.uom_type.value == "timeline":
                return str(c.completion_date) if c.completion_date else ""
            return c.actual_value if c.actual_value is not None else ""

        def status_for(q: Quarter) -> str:
            if q not in by_q:
                return ""
            return by_q[q].goal_status.value

        planned_target = (
            str(goal.target_date)
            if goal.uom_type.value == "timeline" and goal.target_date
            else goal.target_value
            if goal.target_value is not None
            else ""
        )

        rows.append(
            {
                "Employee Name": goal.employee.full_name if goal.employee else "",
                "Department": goal.employee.department if goal.employee else "",
                "Goal Title": goal.title,
                "Thrust Area": goal.thrust_area,
                "UoM Type": goal.uom_type.value,
                "Weightage %": goal.weightage,
                "Planned Target": planned_target,
                "Q1 Actual Achievement": actual_for(Quarter.Q1),
                "Q1 Status": status_for(Quarter.Q1),
                "Q2 Actual Achievement": actual_for(Quarter.Q2),
                "Q2 Status": status_for(Quarter.Q2),
                "Q3 Actual Achievement": actual_for(Quarter.Q3),
                "Q3 Status": status_for(Quarter.Q3),
                "Q4 Actual Achievement": actual_for(Quarter.Q4),
                "Q4 Status": status_for(Quarter.Q4),
                "Goal Status": goal.status.value,
                "Computed Score (latest)": (
                    round(latest_score, 2) if latest_score is not None else ""
                ),
                "Weighted Total": round(weighted_total, 2),
            }
        )
    return rows


def to_csv(rows: list[dict]) -> str:
    if not rows:
        return ""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def to_xlsx(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Achievement"
    if not rows:
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    headers = list(rows[0].keys())
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="2563EB")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    wt_col = headers.index("Weighted Total") + 1
    for row in rows:
        ws.append([row.get(h) for h in headers])
        pct_cell = ws.cell(row=ws.max_row, column=wt_col)
        try:
            val = float(pct_cell.value or 0)
            if val < 0.5:
                pct_cell.fill = PatternFill("solid", fgColor="FCA5A5")
            elif val < 0.8:
                pct_cell.fill = PatternFill("solid", fgColor="FDE68A")
            else:
                pct_cell.fill = PatternFill("solid", fgColor="86EFAC")
        except (TypeError, ValueError):
            pass

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def completion_dashboard(db: AsyncSession, cycle_id: UUID) -> list[dict]:
    cycle = await db.get(PerformanceCycle, cycle_id)
    active_q = get_active_quarter(cycle) if cycle else None

    managers = list(
        (await db.execute(select(User).where(User.role == UserRole.manager))).scalars().all()
    )
    dashboard = []
    for mgr in managers:
        team = list(
            (
                await db.execute(select(User.id).where(User.manager_id == mgr.id))
            ).scalars().all()
        )
        if not team:
            continue
        total = len(team)
        submitted = await db.scalar(
            select(func.count(func.distinct(Goal.employee_id))).where(
                Goal.cycle_id == cycle_id,
                Goal.employee_id.in_(team),
                Goal.status.in_([GoalStatus.submitted, GoalStatus.locked]),
            )
        )
        approved = await db.scalar(
            select(func.count(func.distinct(Goal.employee_id))).where(
                Goal.cycle_id == cycle_id,
                Goal.employee_id.in_(team),
                Goal.status == GoalStatus.locked,
            )
        )

        checkins_pct = 0.0
        if active_q:
            # employees with at least one checkin in the active quarter
            checkin_employees = await db.scalar(
                select(func.count(func.distinct(Goal.employee_id)))
                .join(QuarterlyCheckin, QuarterlyCheckin.goal_id == Goal.id)
                .where(
                    Goal.cycle_id == cycle_id,
                    Goal.employee_id.in_(team),
                    QuarterlyCheckin.quarter == active_q,
                )
            )
            checkins_pct = round((checkin_employees or 0) / total * 100, 1)

        dashboard.append(
            {
                "manager_name": mgr.full_name,
                "team_size": total,
                "goals_submitted_pct": round((submitted or 0) / total * 100, 1),
                "goals_approved_pct": round((approved or 0) / total * 100, 1),
                "checkins_done_pct": checkins_pct,
                "active_quarter": active_q.value if active_q else None,
            }
        )
    return dashboard


async def employee_checkin_matrix(db: AsyncSession, cycle_id: UUID) -> list[dict]:
    """Drill-down view: every employee × every quarter × completion + manager-reviewed flags."""
    employees = list(
        (
            await db.execute(
                select(User).where(User.role == UserRole.employee).order_by(User.full_name)
            )
        )
        .scalars()
        .all()
    )

    rows: list[dict] = []
    for emp in employees:
        manager = await db.get(User, emp.manager_id) if emp.manager_id else None
        goals = list(
            (
                await db.execute(
                    select(Goal).where(
                        Goal.employee_id == emp.id, Goal.cycle_id == cycle_id
                    )
                )
            )
            .scalars()
            .all()
        )
        total_locked = sum(1 for g in goals if g.status == GoalStatus.locked)
        per_quarter: dict[str, dict] = {}
        for qtr in Quarter:
            done_count = 0
            mgr_reviewed = 0
            for g in goals:
                if g.status != GoalStatus.locked:
                    continue
                ck = (
                    await db.execute(
                        select(QuarterlyCheckin).where(
                            QuarterlyCheckin.goal_id == g.id,
                            QuarterlyCheckin.quarter == qtr,
                        )
                    )
                ).scalar_one_or_none()
                if ck:
                    done_count += 1
                    if ck.manager_comment:
                        mgr_reviewed += 1
            per_quarter[qtr.value] = {
                "checkins_done": done_count,
                "of_goals": total_locked,
                "employee_complete": total_locked > 0 and done_count == total_locked,
                "manager_reviewed": mgr_reviewed,
                "manager_complete": total_locked > 0 and mgr_reviewed == total_locked,
            }
        rows.append(
            {
                "employee_id": str(emp.id),
                "employee_name": emp.full_name,
                "department": emp.department,
                "manager_name": manager.full_name if manager else None,
                "total_locked_goals": total_locked,
                "quarters": per_quarter,
            }
        )
    return rows


async def completion_heatmap(db: AsyncSession, cycle_id: UUID) -> dict:
    """Return department x quarter completion rate as a heatmap."""
    departments_res = await db.execute(
        select(User.department).where(User.department.is_not(None)).distinct()
    )
    departments = [d for d in departments_res.scalars().all() if d]
    quarters = [q.value for q in Quarter]

    heatmap = []
    for dept in departments:
        team = list(
            (
                await db.execute(
                    select(User.id).where(User.department == dept, User.role == UserRole.employee)
                )
            )
            .scalars()
            .all()
        )
        if not team:
            continue
        row = {"department": dept}
        for qtr in Quarter:
            done = await db.scalar(
                select(func.count(func.distinct(Goal.employee_id)))
                .join(QuarterlyCheckin, QuarterlyCheckin.goal_id == Goal.id)
                .where(
                    Goal.cycle_id == cycle_id,
                    Goal.employee_id.in_(team),
                    QuarterlyCheckin.quarter == qtr,
                )
            )
            row[qtr.value] = round((done or 0) / len(team) * 100, 1)
        heatmap.append(row)
    return {"departments": [r["department"] for r in heatmap], "quarters": quarters, "rows": heatmap}
